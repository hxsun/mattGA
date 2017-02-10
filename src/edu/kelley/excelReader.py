'''
Created on Jan 27, 2017

@author: haoxsun
'''
from openpyxl import load_workbook
from openpyxl.utils import coordinate_from_string, column_index_from_string
from pymongo import MongoClient
from multiprocessing import Process, Queue

SEC_DOMAIN = "https://www.sec.gov"
# Function to retreive the actual "period of report" and "10k report url"
from lxml import html
import requests
def parseIntroPage(q, url):
    page = requests.get(url)
    tree = html.fromstring(page.content)
    periodOfReportStr = tree.xpath("//div[@id='formDiv']//div[@class='formGrouping'][2]//div[@class='info'][1]/text()")[0]
    actualURLs = tree.xpath("//table[@class='tableFile']//tr[td[4]='10-K']/td[3]/a/@href")
    availableURLs = []
    for url in actualURLs:
        if url.endswith('.txt') or url.endswith('.htm'):
            availableURLs.append(url)
    q.put({'periodOfReport': periodOfReportStr, 'actual10KURL': availableURLs})



import logging
#import os
from time import time
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logging.getLogger('requests').setLevel(logging.CRITICAL)
logger = logging.getLogger(__name__)

def main():
    ts = time()
    fileLocation = "../../../data/raind1A.xlsx"
    # "../../../data/RAIND1.xlsx"
    wb = load_workbook(fileLocation)

    ws = wb.get_sheet_by_name("raind1A")
    
    # "A": COMPANY_FKEY, "B": FILE_DATE, "C": BEST_EDGAR_TICKER, 
    # "D": FORM_FKEY, "E": HTTP_NAME_HTML, "F": PE_DATE,
    # "G": PE_DATE_NUM, "H": FISCAL_YE, "I": GVKEY, "J": CIK
    columns = ["A", "B", "C", "D", "E", "F", "I", "J"]
    row_counter = 0

    # Initiate the mongodb client
    client = MongoClient("localhost", 27017)
    db = client.gadb
    linkCollection = db.linkCollection
    # dataRecords = []
    unMatchedRecordsNum = 0
    q = Queue()
    for row in ws.iter_rows(row_offset=1):

        row_counter = row_counter + 1
        if row[column_index_from_string(columns[0])-1].value is not None:
            companyFKey = row[column_index_from_string(columns[0])-1].value
            fileDate = '{:%Y-%m-%d}'.format(row[column_index_from_string(columns[1])-1].value)
            bestEdgarTicker = row[column_index_from_string(columns[2])-1].value
            formFKey = row[column_index_from_string(columns[3])-1].value
            httpNameHtml = SEC_DOMAIN + row[column_index_from_string(columns[4])-1].value
            peDate = '{:%Y-%m-%d}'.format(row[column_index_from_string(columns[5])-1].value)
            gvKey = row[column_index_from_string(columns[6])-1].value
            cik = row[column_index_from_string(columns[7])-1].value

            qResult = linkCollection.find_one({'peDate':peDate, 'companyFKey': companyFKey})
            if (qResult is not None and 'periodOfReport' in qResult):
                continue
            #results = parseIntroPage(url)
            p = Process(target=parseIntroPage, args=(q, httpNameHtml))
            p.start()
            results = q.get()
            p.join()
            dataDic = {}
            dataDic['companyFKey'] = companyFKey
            dataDic['fileDate'] = fileDate
            dataDic['bestEdgarTicker'] = bestEdgarTicker
            dataDic['formFKey'] = formFKey
            dataDic['httpNameHtml'] = httpNameHtml
            dataDic['peDate'] = peDate
            dataDic['gvKey'] = gvKey
            dataDic['CIK'] = cik
            # after getting the results from the page crawling process then update the table with the record.
            # Check the the actual date in 10k is same as the date in excel file.
            matched = True
            if (results['periodOfReport'] != peDate):
                unMatchedRecordsNum = unMatchedRecordsNum + 1
                matched = False
                print("UnMatched, origDate:%s, newDate:%s, companyFKey:%s"%
                      (peDate, results['periodOfReport'], companyFKey))
            dataDic['Matched'] = matched
            dataDic['periodOfReport'] = results['periodOfReport']
            # 10KURL is array of the urls ending with txt or htm
            dataDic['10KURL'] = results['actual10KURL']
            # Update or insert the record to mongodb
            linkCollection.update({'peDate':peDate, 'companyFKey': companyFKey}, 
                                  {"$set": dataDic}, upsert = True)
            
            #print(ws[columns[0] + row] + " " + ws[columns[1] + row] + " "+ ws[columns[2] + row])
    client.close()
    print("This sheet has %d rows" % (row_counter))
    print("There are %d records having unmatched urls." % (unMatchedRecordsNum))
    print('Took {}s'.format(time() - ts))
   
if __name__ == '__main__':
    main()

